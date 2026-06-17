from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Iterable
from urllib.parse import urlsplit, urlunsplit

from openpyxl import load_workbook

from src.models import ReferenceLink


_WHITESPACE_RE = re.compile(r"\s+")
_ALIAS_SPLIT_RE = re.compile(r"[;\n,|]+")
_URL_RE = re.compile(r"https?://\S+", re.IGNORECASE)


def normalize_text(value: object) -> str:
    text = "" if value is None else str(value)
    return _WHITESPACE_RE.sub(" ", text).strip()


def normalize_domain(value: object) -> str:
    text = normalize_text(value).lower()
    if not text:
        return ""

    url_match = _URL_RE.search(text)
    candidate = url_match.group(0) if url_match is not None else text
    parsed = urlsplit(candidate if "://" in candidate else f"https://{candidate}")
    host = parsed.netloc or parsed.path
    host = host.split("/")[0]
    if host.startswith("www."):
        host = host[4:]
    return host.strip().lower()


def normalize_page_url(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""

    url_match = _URL_RE.search(text)
    candidate = url_match.group(0) if url_match is not None else text
    parsed = urlsplit(candidate if "://" in candidate else f"https://{candidate}")
    scheme = parsed.scheme.lower() if parsed.scheme else "https"
    netloc = parsed.netloc.lower()
    path = parsed.path.rstrip("/") or parsed.path
    return urlunsplit((scheme, netloc, path, parsed.query, ""))


def normalize_tariff_name(value: object) -> str:
    return normalize_text(value).lower()


def split_aliases(value: object) -> tuple[str, ...]:
    text = normalize_text(value)
    if not text:
        return ()
    aliases = [normalize_text(part) for part in _ALIAS_SPLIT_RE.split(text)]
    return tuple(alias for alias in aliases if alias)


def parse_bool(value: object, default: bool = True) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    text = normalize_text(value).lower()
    if not text:
        return default
    if text in {"1", "true", "yes", "y", "да", "on"}:
        return True
    if text in {"0", "false", "no", "n", "нет", "off"}:
        return False
    return default


def _normalize_header(value: object) -> str:
    return normalize_text(value).lower()


def _header_map(headers: Iterable[object]) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, header in enumerate(headers):
        key = _normalize_header(header)
        if key and key not in result:
            result[key] = index
    return result


def _first_header_index(header_map: dict[str, int], aliases: tuple[str, ...]) -> int | None:
    for alias in aliases:
        index = header_map.get(_normalize_header(alias))
        if index is not None:
            return index
    return None


def _row_has_content(row: tuple[object, ...]) -> bool:
    return any(cell is not None and normalize_text(cell) for cell in row)


def _looks_like_site_context_row(row: tuple[object, ...]) -> bool:
    first = normalize_text(row[0]) if row else ""
    if not first:
        return False
    trailing = [normalize_text(cell) for cell in row[1:]]
    if any(trailing):
        return False
    return first.startswith(("http://", "https://")) or "." in first


def _load_xlsx_rows(path: Path) -> list[tuple[object, ...]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    return [tuple(row) for row in sheet.iter_rows(values_only=True)]


def _load_csv_rows(path: Path) -> list[tuple[object, ...]]:
    encodings = ("utf-8-sig", "utf-8", "cp1251")
    raw_text = None
    for encoding in encodings:
        try:
            raw_text = path.read_text(encoding=encoding)
            break
        except UnicodeDecodeError:
            continue
    if raw_text is None:
        raw_text = path.read_text(encoding="utf-8", errors="replace")

    sample = raw_text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.get_dialect("excel")

    reader = csv.reader(raw_text.splitlines(), dialect)
    return [tuple(row) for row in reader]


def _build_reference_link(
    header_map: dict[str, int],
    row: tuple[object, ...],
    row_number: int,
) -> ReferenceLink | None:
    def get_cell(*aliases: str, default: object = "") -> object:
        index = _first_header_index(header_map, aliases)
        if index is None or index >= len(row):
            return default
        value = row[index]
        return default if value is None else value

    active = parse_bool(get_cell("Активен", "Active"), default=True)
    if not active:
        return None

    domain = normalize_domain(get_cell("Домен", "Domain"))
    page_url = normalize_page_url(get_cell("URL проверяемой страницы", "Page URL", "URL", "page_url"))
    tariff_name = normalize_tariff_name(get_cell("Название тарифа", "Tariff Name", "tariff_name"))
    expected_url_part = normalize_text(
        get_cell("Что сверяем", "Эталонная часть ссылки", "Expected URL Part", "Expected Part", "expected_url_part")
    )
    match_type = normalize_text(get_cell("Тип сравнения", "Match Type", "match_type")) or "contains"
    comment = normalize_text(get_cell("Комментарий", "Comment", "comment"))
    aliases = split_aliases(get_cell("Алиасы тарифа", "Tariff Aliases", "Aliases", "aliases"))

    if not domain or not tariff_name or not expected_url_part:
        return None
    if match_type not in {"contains", "regex"}:
        return None

    return ReferenceLink(
        domain=domain,
        page_url=page_url,
        tariff_name=tariff_name,
        expected_url_part=expected_url_part,
        match_type=match_type,
        comment=comment,
        active=active,
        aliases=aliases,
    )


def _load_compact_rows(rows: list[tuple[object, ...]]) -> list[ReferenceLink]:
    parsed: list[ReferenceLink] = []
    current_domain = ""

    for row_number, row in enumerate(rows[1:], start=2):
        if not _row_has_content(row):
            continue

        if _looks_like_site_context_row(row):
            current_domain = normalize_domain(row[0])
            continue

        tariff_name = normalize_tariff_name(row[0] if len(row) > 0 else "")
        if not tariff_name:
            continue

        expected_url_part = normalize_text(row[3] if len(row) > 3 else "")
        if not expected_url_part:
            continue

        comment_parts = []
        primary_link = normalize_text(row[1] if len(row) > 1 else "")
        opened_as = normalize_text(row[2] if len(row) > 2 else "")
        if primary_link:
            comment_parts.append(f"primary={primary_link}")
        if opened_as:
            comment_parts.append(f"opened_as={opened_as}")
        comment = "; ".join(comment_parts)

        if not current_domain:
            continue

        parsed.append(
            ReferenceLink(
                domain=current_domain,
                page_url="",
                tariff_name=tariff_name,
                expected_url_part=expected_url_part,
                match_type="contains",
                comment=comment,
                active=True,
                aliases=(),
            )
        )

    return parsed


def load_reference_links(reference_file: str | Path) -> list[ReferenceLink]:
    path = Path(reference_file)
    if not path.exists():
        raise FileNotFoundError(f"Reference file not found: {path}")

    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        rows = _load_xlsx_rows(path)
    elif suffix == ".csv":
        rows = _load_csv_rows(path)
    else:
        raise ValueError(f"Unsupported reference file format: {path.suffix}")

    if not rows:
        return []

    header_names = {_normalize_header(header) for header in rows[0] if _normalize_header(header)}
    compact_headers = {"сайт", "ссылка первичная", "как открывается", "что сверяем"}
    if compact_headers.issubset(header_names):
        parsed = _load_compact_rows(rows)
        if not parsed:
            raise ValueError(f"No valid reference rows found in compact workbook: {path}")
        return parsed

    header_map = _header_map(rows[0])
    parsed: list[ReferenceLink] = []
    for row_number, row in enumerate(rows[1:], start=2):
        if not _row_has_content(row):
            continue
        reference = _build_reference_link(header_map, row, row_number)
        if reference is not None:
            parsed.append(reference)

    if not parsed:
        raise ValueError(f"No valid reference rows found in workbook: {path}")

    return parsed
