from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.models import (
    ReportRow,
    VALIDATION_STATUS_NO_FACTUAL_LINK,
    VALIDATION_STATUS_NO_REFERENCE,
    VALIDATION_STATUS_NOT_OK,
    VALIDATION_STATUS_OK,
)


HEADERS = [
    "Домен",
    "URL проверяемой страницы",
    "Название тарифа",
    "Ссылка после клика",
    "Статус загрузки страницы",
    "HTTP-статус",
    "Финальный URL",
    "Ошибка",
    "Дата и время проверки",
    "Оператор",
    "Номер карточки",
    "Тип перехода",
    "Исходный href",
    "Время загрузки, мс",
    "Окружение запуска",
    "Признак продуктовой ошибки",
]

ERROR_FILL = PatternFill(fill_type="solid", fgColor="FDE2E1")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
OK_FILL = PatternFill(fill_type="solid", fgColor="E2F6E9")
NOT_OK_FILL = PatternFill(fill_type="solid", fgColor="FDE2E1")
NO_REFERENCE_FILL = PatternFill(fill_type="solid", fgColor="FFF4CC")
NO_FACTUAL_LINK_FILL = PatternFill(fill_type="solid", fgColor="FCE5CD")


def _as_cell(value: object) -> object:
    if isinstance(value, bool):
        return "true" if value else "false"
    return value


def build_report_path(report_dir: str | Path, timestamp: datetime | None = None) -> Path:
    stamp = (timestamp or datetime.now()).strftime("%Y-%m-%d_%H-%M")
    return Path(report_dir) / f"partner_links_mobile_{stamp}.xlsx"


def write_report(
    rows: Iterable[ReportRow],
    report_dir: str | Path,
    timestamp: datetime | None = None,
) -> Path:
    report_path = build_report_path(report_dir, timestamp=timestamp)
    report_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "partner_links_mobile"
    sheet.freeze_panes = "A2"

    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in rows:
        values = [
            row.domain,
            row.checked_page_url,
            row.tariff_name,
            row.click_url,
            row.page_load_status,
            row.http_status,
            row.final_url,
            row.error,
            row.checked_at,
            row.operator,
            row.card_number,
            row.transition_type,
            row.source_href,
            row.load_ms,
            row.environment,
            row.product_error,
        ]
        sheet.append([_as_cell(value) for value in values])
        if row.product_error:
            for cell in sheet[sheet.max_row]:
                cell.fill = ERROR_FILL

    for column_cells in sheet.columns:
        values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
        width = min(max(len(value) for value in values) + 2, 60)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    workbook.save(report_path)
    return report_path


VALIDATION_HEADERS = HEADERS + [
    "Эталонная часть ссылки",
    "Итог проверки партнерской ссылки",
    "Ошибка проверки партнерской ссылки",
    "Ключ сопоставления",
    "Тип сравнения",
]


def _validation_row_fill(status: str) -> PatternFill | None:
    if status == VALIDATION_STATUS_OK:
        return OK_FILL
    if status == VALIDATION_STATUS_NOT_OK:
        return NOT_OK_FILL
    if status == VALIDATION_STATUS_NO_REFERENCE:
        return NO_REFERENCE_FILL
    if status == VALIDATION_STATUS_NO_FACTUAL_LINK:
        return NO_FACTUAL_LINK_FILL
    return None


def write_validation_report(
    rows: Iterable[ReportRow],
    report_path: str | Path,
    timestamp: datetime | None = None,
) -> Path:
    path = Path(report_path)
    if path.suffix.lower() != ".xlsx":
        path = path / f"partner_links_mobile_validated_{(timestamp or datetime.now()).strftime('%Y-%m-%d_%H-%M')}.xlsx"
    report_path = path
    report_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "partner_links_mobile_validated"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(VALIDATION_HEADERS))}1"

    sheet.append(VALIDATION_HEADERS)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in rows:
        values = [
            row.domain,
            row.checked_page_url,
            row.tariff_name,
            row.click_url,
            row.page_load_status,
            row.http_status,
            row.final_url,
            row.error,
            row.checked_at,
            row.operator,
            row.card_number,
            row.transition_type,
            row.source_href,
            row.load_ms,
            row.environment,
            row.product_error,
            row.reference_part,
            row.validation_status,
            row.validation_error,
            row.match_key,
            row.comparison_type,
        ]
        sheet.append([_as_cell(value) for value in values])
        fill = _validation_row_fill(row.validation_status)
        if fill is not None:
            for cell in sheet[sheet.max_row]:
                cell.fill = fill

    for column_cells in sheet.columns:
        values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
        width = min(max(len(value) for value in values) + 2, 60)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    workbook.save(report_path)
    return report_path
