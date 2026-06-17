from __future__ import annotations

import argparse
import os
from pathlib import Path

from openpyxl import load_workbook

from src.alert_sender import build_validation_alert_message, send_validation_alert
from src.link_matcher import ReferenceIndex, evaluate_validation
from src.reference_loader import normalize_text, parse_bool, load_reference_links
from src.models import ReportRow
from src.report_writer import write_validation_report
from src.result_evaluator import build_validation_release_failure_message, summarize_validation_rows


INPUT_COLUMN_ALIASES = {
    "domain": ("Домен", "Domain", "domain"),
    "checked_page_url": ("URL проверяемой страницы", "Page URL", "page_url"),
    "tariff_name": ("Название тарифа", "Tariff Name", "tariff_name"),
    "click_url": ("Ссылка после клика", "Clicked URL", "click_url"),
    "page_load_status": ("Статус загрузки страницы", "Page load status", "page_load_status"),
    "http_status": ("HTTP-статус", "HTTP status", "http_status"),
    "final_url": ("Финальный URL", "Final URL", "final_url"),
    "error": ("Ошибка", "Error", "error"),
    "checked_at": ("Дата и время проверки", "Checked at", "checked_at"),
    "operator": ("Оператор", "Operator", "operator"),
    "card_number": ("Номер карточки", "Card number", "card_number"),
    "transition_type": ("Тип перехода", "Transition type", "transition_type"),
    "source_href": ("Исходный href", "Source href", "source_href"),
    "load_ms": ("Время загрузки, мс", "Load ms", "load_ms"),
    "environment": ("Окружение запуска", "Environment", "environment"),
    "product_error": ("Признак продуктовой ошибки", "Product error", "product_error"),
}


def _normalize_header(value: object) -> str:
    return normalize_text(value).lower()


def _header_map(headers: list[object]) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, header in enumerate(headers):
        key = _normalize_header(header)
        if key and key not in result:
            result[key] = index
    return result


def _first_index(header_map: dict[str, int], aliases: tuple[str, ...]) -> int | None:
    for alias in aliases:
        index = header_map.get(_normalize_header(alias))
        if index is not None:
            return index
    return None


def _get_cell(row: tuple[object, ...], header_map: dict[str, int], *aliases: str, default: object = "") -> object:
    index = _first_index(header_map, aliases)
    if index is None or index >= len(row):
        return default
    value = row[index]
    return default if value is None else value


def _latest_report_in(directory: Path) -> Path | None:
    reports = [
        path
        for path in directory.glob("partner_links_mobile_*.xlsx")
        if path.is_file() and not path.name.startswith("partner_links_mobile_validated_")
    ]
    if not reports:
        return None
    return max(reports, key=lambda path: path.stat().st_mtime)


def _resolve_input_report_path(input_report: str | Path | None) -> Path:
    if input_report:
        path = Path(input_report)
        if path.exists():
            return path

        if not path.is_absolute() and path.parts[:1] != ("reports",):
            reports_path = Path("reports") / path
            if reports_path.exists():
                return reports_path

        raise FileNotFoundError(f"Input report not found: {path}")

    latest_report = _latest_report_in(Path("reports"))
    if latest_report is not None:
        return latest_report

    latest_report = _latest_report_in(Path.cwd())
    if latest_report is not None:
        return latest_report

    raise FileNotFoundError("No partner_links_mobile report found")


def load_input_rows(input_report: str | Path) -> list[ReportRow]:
    path = _resolve_input_report_path(input_report)
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"Unsupported input report format: {path.suffix}")

    workbook = load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = [tuple(row) for row in sheet.iter_rows(values_only=True)]
    if not rows:
        return []

    header_map = _header_map(list(rows[0]))
    report_rows = []
    for row in rows[1:]:
        if not any(normalize_text(cell) for cell in row if cell is not None):
            continue
        report_rows.append(
            ReportRow(
                domain=normalize_text(_get_cell(row, header_map, *INPUT_COLUMN_ALIASES["domain"])),
                checked_page_url=normalize_text(_get_cell(row, header_map, *INPUT_COLUMN_ALIASES["checked_page_url"])),
                tariff_name=normalize_text(_get_cell(row, header_map, *INPUT_COLUMN_ALIASES["tariff_name"])),
                click_url=normalize_text(_get_cell(row, header_map, *INPUT_COLUMN_ALIASES["click_url"])),
                page_load_status=normalize_text(_get_cell(row, header_map, *INPUT_COLUMN_ALIASES["page_load_status"])),
                http_status=normalize_text(_get_cell(row, header_map, *INPUT_COLUMN_ALIASES["http_status"])),
                final_url=normalize_text(_get_cell(row, header_map, *INPUT_COLUMN_ALIASES["final_url"])),
                error=normalize_text(_get_cell(row, header_map, *INPUT_COLUMN_ALIASES["error"])),
                checked_at=normalize_text(_get_cell(row, header_map, *INPUT_COLUMN_ALIASES["checked_at"])),
                operator=normalize_text(_get_cell(row, header_map, *INPUT_COLUMN_ALIASES["operator"])),
                card_number=normalize_text(_get_cell(row, header_map, *INPUT_COLUMN_ALIASES["card_number"])),
                transition_type=normalize_text(_get_cell(row, header_map, *INPUT_COLUMN_ALIASES["transition_type"])),
                source_href=normalize_text(_get_cell(row, header_map, *INPUT_COLUMN_ALIASES["source_href"])),
                load_ms=normalize_text(_get_cell(row, header_map, *INPUT_COLUMN_ALIASES["load_ms"])),
                environment=normalize_text(_get_cell(row, header_map, *INPUT_COLUMN_ALIASES["environment"])),
                product_error=parse_bool(
                    _get_cell(row, header_map, *INPUT_COLUMN_ALIASES["product_error"]),
                    default=False,
                ),
            )
        )

    return report_rows


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Validate partner links against the reference file")
    parser.add_argument(
        "--input-report",
        default="",
        help="Path to the first iteration xlsx report. If omitted, the latest report from reports/ is used.",
    )
    parser.add_argument("--reference-file", required=True, help="Path to the Links_mobile_tarriffs file")
    parser.add_argument("--output-report", required=True, help="Path to the validated xlsx report")
    parser.add_argument("--run-mode", default="pilot", choices=("pilot", "release"))
    parser.add_argument(
        "--use-final-url-as-fallback",
        action="store_true",
        default=False,
        help="Use final URL if clicked URL is empty",
    )
    parser.add_argument(
        "--send-alert",
        action="store_true",
        default=bool(os.environ.get("JENKINS_HOME")),
        help="Send Telegram alert through proxy",
    )
    return parser


def run(argv: list[str] | None = None) -> int:
    parser = _build_parser()
    args = parser.parse_args(argv)

    input_rows = load_input_rows(args.input_report)
    reference_links = load_reference_links(args.reference_file)
    reference_index = ReferenceIndex.build(reference_links)

    validated_rows = []
    for row in input_rows:
        result = evaluate_validation(row, reference_index, use_final_url_as_fallback=args.use_final_url_as_fallback)
        validated_rows.append(result.row)

    output_path = write_validation_report(validated_rows, args.output_report)
    summary = summarize_validation_rows(validated_rows, str(output_path), args.run_mode)

    checked_at = next((row.checked_at for row in validated_rows if row.checked_at), "")
    build_url = (os.getenv("BUILD_URL") or "").strip() or None
    alert_message = build_validation_alert_message(
        summary,
        checked_at=checked_at,
        build_url=build_url,
        report_path=str(output_path),
    )

    print(
        f"[SUMMARY] total={summary.total_rows} ok={summary.ok_rows} not_ok={summary.not_ok_rows} "
        f"no_reference={summary.no_reference_rows} no_factual_link={summary.no_factual_link_rows} "
        f"report={summary.report_path}"
    )

    if args.send_alert:
        alert_result = send_validation_alert(alert_message)
        print(f"[ALERT] status={alert_result.status}")
    else:
        print("[ALERT] skipped by parameter")

    if args.run_mode == "release" and summary.product_error_rows > 0:
        raise SystemExit(build_validation_release_failure_message(summary))

    return 0


def main() -> None:
    raise SystemExit(run())


if __name__ == "__main__":
    main()
