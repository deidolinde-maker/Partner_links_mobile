from __future__ import annotations

from dataclasses import replace
from pathlib import Path

import json

import pytest
from openpyxl import Workbook, load_workbook

from src.alert_sender import build_validation_alert_message, send_validation_alert
from src.link_matcher import ReferenceIndex, evaluate_validation
from src.models import (
    ReportRow,
    VALIDATION_STATUS_NO_FACTUAL_LINK,
    VALIDATION_STATUS_NO_REFERENCE,
    VALIDATION_STATUS_NOT_OK,
    VALIDATION_STATUS_OK,
)
from src.reference_loader import load_reference_links
from src.report_writer import write_validation_report
from src.result_evaluator import summarize_validation_rows
from src.validate_partner_links import load_input_rows


REFERENCE_HEADERS = [
    "Домен",
    "URL проверяемой страницы",
    "Название тарифа",
    "Что сверяем",
    "Тип сравнения",
    "Алиасы тарифа",
    "Активен",
]


def _write_xlsx(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def _build_input_row(**overrides: object) -> ReportRow:
    base = dict(
        domain="t2-ru.online",
        checked_page_url="https://t2-ru.online/mobilnaya-svyaz",
        tariff_name="МОЙ ОНЛАЙН",
        click_url="https://krasnodar.t2.ru/tariffs?utm_campaign=tariffs_webdealer_ooo_online_services_t2-ru_online&utm_medium=t2-ru_online&utm_source=webdealer&pageParams=askForRegion%3Dtrue",
        page_load_status="Успешно",
        http_status="200",
        final_url="https://krasnodar.t2.ru/tariffs?utm_campaign=tariffs_webdealer_ooo_online_services_t2-ru_online&utm_medium=t2-ru_online&utm_source=webdealer&pageParams=askForRegion%3Dtrue",
        error="",
        checked_at="2026-06-11 09:55:23",
        operator="T2",
        card_number="6",
        transition_type="new_tab",
        source_href="https://t2.ru/tariffs?utm_source=webdealer&utm_medium=t2-ru_online&utm_campaign=tariffs_webdealer_ooo_online_services_t2-ru_online",
        load_ms="12943",
        environment="run_mode=pilot; target=domain; browser=headless; trace=off; screenshot=off",
        product_error=False,
    )
    base.update(overrides)
    return ReportRow(**base)


def test_reference_loader_supports_aliases_and_inactive_rows(tmp_path: Path) -> None:
    reference_path = tmp_path / "Links_mobile_tarriffs.xlsx"
    _write_xlsx(
        reference_path,
        REFERENCE_HEADERS,
        [
            [
                "t2-ru.online",
                "https://t2-ru.online/mobilnaya-svyaz",
                "Мой Онлайн",
                "krasnodar.t2.ru/tariffs",
                "contains",
                "МОЙ ОНЛАЙН+;МОЙ ОНЛАЙН ПЛЮС",
                True,
            ],
            [
                "t2-ru.online",
                "https://t2-ru.online/mobilnaya-svyaz",
                "BLACK",
                "black",
                "contains",
                "",
                False,
            ],
        ],
    )

    references = load_reference_links(reference_path)
    index = ReferenceIndex.build(references)
    result = evaluate_validation(
        _build_input_row(tariff_name="МОЙ ОНЛАЙН+"),
        index,
    )

    assert result.status == VALIDATION_STATUS_OK
    assert result.reference_part == "krasnodar.t2.ru/tariffs"
    assert result.match_key == "t2-ru.online::/mobilnaya-svyaz::мой онлайн+"


def test_reference_loader_supports_compact_site_table_format(tmp_path: Path) -> None:
    reference_path = tmp_path / "Ссылки для парсера.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Лист1"
    sheet.append(["Сайт", "Ссылка первичная", "Как открывается", "Что сверяем", None])
    sheet.append(["https://mts-home.online/", None, None, None, None])
    sheet.append(
        [
            "МТС Супер",
            "https://tracking.pn.mts.ru/tracker?channelId=991de3b9-6956-40d7-8c09-f0eaa3946243&lid=166014",
            "https://spb.mts.ru/personal/mobilnaya-svyaz/tarifi/vse-tarifi/mts-super/?utm_source=mtspn",
            "mts.ru/personal/mobilnaya-svyaz/tarifi/vse-tarifi/mts-super/?utm_source=mtspn",
            "1. Может быть разный регион",
        ]
    )
    sheet.append(
        [
            "РИИЛ",
            "https://tracking.pn.mts.ru/tracker?channelId=991de3b9-6956-40d7-8c09-f0eaa3946243&lid=368458",
            "https://moskva.mts.ru/personal/mobilnaya-svyaz/tarifi/vse-tarifi/riil/?utm_source=mtspn",
            "mts.ru/personal/mobilnaya-svyaz/tarifi/vse-tarifi/riil/?utm_source=mtspn",
            None,
        ]
    )
    workbook.save(reference_path)

    references = load_reference_links(reference_path)
    index = ReferenceIndex.build(references)

    result = evaluate_validation(
        _build_input_row(
            domain="mts-home.online",
            checked_page_url="https://mts-home.online/",
            tariff_name="МТС Супер",
            click_url="https://spb.mts.ru/personal/mobilnaya-svyaz/tarifi/vse-tarifi/mts-super/?utm_source=mtspn",
            final_url="https://spb.mts.ru/personal/mobilnaya-svyaz/tarifi/vse-tarifi/mts-super/?utm_source=mtspn",
        ),
        index,
    )

    assert result.status == VALIDATION_STATUS_OK
    assert result.match_key == "mts-home.online::\u0441\u0443\u043f\u0435\u0440"
    assert result.reference_part == "mts.ru/personal/mobilnaya-svyaz/tarifi/vse-tarifi/mts-super/?utm_source=mtspn"


def test_reference_loader_extracts_domain_from_labelled_site_row(tmp_path: Path) -> None:
    reference_path = tmp_path / "Ссылки для парсера.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Сайт", "Ссылка первичная", "Как открывается", "Что сверяем"])
    sheet.append(["Билайн - https://beeline.ru/", None, None, None])
    sheet.append(
        [
            "Все тарифы",
            "https://tracking.example/link",
            "https://beeline.ru/customers/products/toptariffs/?utm_source=mobideal&utm_medium=cpa&utm_campaign=landing",
            "beeline.ru/customers/products/toptariffs/?utm_source=mobideal&utm_medium=cpa&utm_campaign=landing",
        ]
    )
    workbook.save(reference_path)

    references = load_reference_links(reference_path)

    assert references[0].domain == "beeline.ru"
    assert references[0].tariff_name == "все тарифы"


def test_reference_loader_normalizes_tariff_names_and_comments(tmp_path: Path) -> None:
    reference_path = tmp_path / "Links_mobile_tarriffs.xlsx"
    _write_xlsx(
        reference_path,
        REFERENCE_HEADERS,
        [
            [
                "mts-home.online",
                "https://mts-home.online/",
                "\u041c\u0422\u0421 Junior",
                "mts.ru/personal/mobilnaya-svyaz/uslugi/mobilnaya-svyaz/mts-junior/?utm_source=mtspn",
                "contains",
                "",
                True,
            ],
            [
                "t2-ru.online",
                "https://t2-ru.online/mobilnaya-svyaz",
                "\u041f\u0410\u0420\u0422\u041d\u0415\u0420 \u041c (\u043d\u0435\u0442 \u043d\u0430 \u043e\u0431\u0449\u0435\u0439 \u0441\u0442\u0440\u0430\u043d\u0438\u0446\u0435)",
                "t2.ru/promo/partner-all?utm_source=webdealer",
                "contains",
                "",
                True,
            ],
        ],
    )

    references = load_reference_links(reference_path)

    assert references[0].tariff_name == "junior"
    assert references[1].tariff_name == "\u043f\u0430\u0440\u0442\u043d\u0435\u0440 \u043c"
    assert "\u043d\u0435\u0442 \u043d\u0430 \u043e\u0431\u0449\u0435\u0439 \u0441\u0442\u0440\u0430\u043d\u0438\u0446\u0435" in references[1].comment.lower()


def test_reference_index_ignores_region_in_page_url(tmp_path: Path) -> None:
    reference_path = tmp_path / "Links_mobile_tarriffs.xlsx"
    _write_xlsx(
        reference_path,
        REFERENCE_HEADERS,
        [
            [
                "beeline.ru",
                "https://spb.beeline.ru/customers/products/toptariffs/",
                "Топ тариф",
                "beeline.ru/customers/products/toptariffs/?utm_source=mobideal&utm_medium=cpa&utm_campaign=landing",
                "contains",
                "",
                True,
            ],
        ],
    )

    references = load_reference_links(reference_path)
    index = ReferenceIndex.build(references)

    result = evaluate_validation(
        _build_input_row(
            domain="beeline.ru",
            checked_page_url="https://moskva.beeline.ru/customers/products/toptariffs/",
            tariff_name="Топ тариф",
            click_url="https://moskva.beeline.ru/customers/products/toptariffs/?utm_source=mobideal&utm_medium=cpa&utm_campaign=landing",
            final_url="https://moskva.beeline.ru/customers/products/toptariffs/?utm_source=mobideal&utm_medium=cpa&utm_campaign=landing",
        ),
        index,
    )

    assert result.status == VALIDATION_STATUS_OK
    assert result.match_key == "beeline.ru::/customers/products/toptariffs::топ тариф"


def test_reference_index_normalizes_mts_home_region_subdomain(tmp_path: Path) -> None:
    reference_path = tmp_path / "Links_mobile_tarriffs.xlsx"
    _write_xlsx(
        reference_path,
        REFERENCE_HEADERS,
        [
            [
                "mts-home.online",
                "https://mts-home.online/",
                "МТС Junior",
                "mts.ru/personal/mobilnaya-svyaz/uslugi/mobilnaya-svyaz/mts-junior/?utm_source=mtspn",
                "contains",
                "",
                True,
            ],
        ],
    )

    references = load_reference_links(reference_path)
    index = ReferenceIndex.build(references)

    result = evaluate_validation(
        _build_input_row(
            domain="sankt-peterburg.mts-home.online",
            checked_page_url="https://sankt-peterburg.mts-home.online/mobilnaya-svyaz",
            tariff_name="МТС Junior",
            click_url="https://spb.mts.ru/personal/mobilnaya-svyaz/uslugi/mobilnaya-svyaz/mts-junior/?utm_source=mtspn",
            final_url="https://spb.mts.ru/personal/mobilnaya-svyaz/uslugi/mobilnaya-svyaz/mts-junior/?utm_source=mtspn",
        ),
        index,
    )

    assert result.status == VALIDATION_STATUS_OK
    assert result.match_key == "mts-home.online::junior"


def test_validation_ignores_landing_number_in_actual_url(tmp_path: Path) -> None:
    reference_path = tmp_path / "Links_mobile_tarriffs.xlsx"
    _write_xlsx(
        reference_path,
        REFERENCE_HEADERS,
        [
            [
                "mts-home.online",
                "https://mts-home.online/",
                "Junior",
                "mts.ru/personal/mobilnaya-svyaz/uslugi/mobilnaya-svyaz/mts-junior/?utm_source=mtspn&utm_medium=cpa&utm_content=&oid=841483&utm_campaign=c_mtspn_s_101internet_r_rf_f_mix_pl_Лендинг 201_t_cpa_a_broad_k_promo&clickid",
                "contains",
                "",
                True,
            ],
        ],
    )

    references = load_reference_links(reference_path)
    index = ReferenceIndex.build(references)

    result = evaluate_validation(
        _build_input_row(
            domain="mts-home.online",
            checked_page_url="https://mts-home.online/",
            tariff_name="Junior",
            click_url="https://moskva.mts.ru/personal/mobilnaya-svyaz/uslugi/mobilnaya-svyaz/mts-junior/?utm_source=mtspn&utm_medium=cpa&utm_content=&oid=841483&utm_campaign=c_mtspn_s_101internet_r_rf_f_mix_pl_Лендинг 202_t_cpa_a_broad_k_promo&clickid=b351b0ba-9038-47f7-92bd-d5c45c83ab30&erid=",
            final_url="https://moskva.mts.ru/personal/mobilnaya-svyaz/uslugi/mobilnaya-svyaz/mts-junior/?utm_source=mtspn&utm_medium=cpa&utm_content=&oid=841483&utm_campaign=c_mtspn_s_101internet_r_rf_f_mix_pl_Лендинг 202_t_cpa_a_broad_k_promo&clickid=b351b0ba-9038-47f7-92bd-d5c45c83ab30&erid=",
        ),
        index,
    )

    assert result.status == VALIDATION_STATUS_OK
    assert result.reference_part == "mts.ru/personal/mobilnaya-svyaz/uslugi/mobilnaya-svyaz/mts-junior/?utm_source=mtspn&utm_medium=cpa&utm_content=&oid=841483&utm_campaign=c_mtspn_s_101internet_r_rf_f_mix_pl_Лендинг 201_t_cpa_a_broad_k_promo&clickid"


def test_reference_index_uses_general_page_row_for_any_tariff(tmp_path: Path) -> None:
    reference_path = tmp_path / "Links_mobile_tarriffs.xlsx"
    _write_xlsx(
        reference_path,
        REFERENCE_HEADERS,
        [
            [
                "online-beeline.ru",
                "https://online-beeline.ru/",
                "\u0412\u0441\u0435 \u0442\u0430\u0440\u0438\u0444\u044b (\u043e\u0431\u0449\u0430\u044f \u0441\u0442\u0440\u0430\u043d\u0438\u0446\u0430)",
                "beeline.ru/customers/products/toptariffs/?utm_source=mobideal&utm_medium=cpa&utm_campaign=landing",
                "contains",
                "",
                True,
            ],
        ],
    )

    references = load_reference_links(reference_path)
    index = ReferenceIndex.build(references)

    result = evaluate_validation(
        _build_input_row(
            domain="beeline-internet.online",
            checked_page_url="https://beeline-internet.online/tariffs-mobile",
            tariff_name="bee HIT",
            click_url="https://moskva.beeline.ru/customers/products/toptariffs/?utm_source=mobideal&utm_medium=cpa&utm_campaign=landing",
            final_url="https://moskva.beeline.ru/customers/products/toptariffs/?utm_source=mobideal&utm_medium=cpa&utm_campaign=landing",
        ),
        index,
    )

    assert result.status == VALIDATION_STATUS_OK
    assert result.match_key == "beeline.ru::*::\u0432\u0441\u0435 \u0442\u0430\u0440\u0438\u0444\u044b"
    assert result.reference_part == "beeline.ru/customers/products/toptariffs/?utm_source=mobideal&utm_medium=cpa&utm_campaign=landing"


def test_reference_index_prefers_exact_tariff_over_general_page_row(tmp_path: Path) -> None:
    reference_path = tmp_path / "Links_mobile_tarriffs.xlsx"
    _write_xlsx(
        reference_path,
        REFERENCE_HEADERS,
        [
            [
                "t2-ru.online",
                "",
                "\u0412\u0441\u0435 \u0442\u0430\u0440\u0438\u0444\u044b (\u043e\u0431\u0449\u0430\u044f \u0441\u0442\u0440\u0430\u043d\u0438\u0446\u0430)",
                "t2.ru/tariffs?utm_campaign=tariffs_webdealer_ooo_online_services_t2-ru_online&utm_medium=t2-ru_online&utm_source=webdealer&pageParams=askForRegion%3Dtrue",
                "contains",
                "",
                True,
            ],
            [
                "t2-ru.online",
                "",
                "\u041f\u0410\u0420\u0422\u041d\u0415\u0420 \u041c",
                "t2.ru/promo/partner-all?utm_campaign=partner_m_webdealer_ooo_online_services_t2-ru_online&utm_medium=t2-ru_online&utm_source=webdealer&pageParams=askForRegion%3Dtrue",
                "contains",
                "",
                True,
            ],
        ],
    )

    references = load_reference_links(reference_path)
    index = ReferenceIndex.build(references)

    result = evaluate_validation(
        _build_input_row(
            domain="t2-ru.online",
            checked_page_url="https://t2-ru.online/mobilnaya-svyaz",
            tariff_name="\u041f\u0410\u0420\u0422\u041d\u0415\u0420 \u041c",
            click_url="https://t2.ru/promo/partner-all?utm_source=webdealer&utm_medium=t2-ru_online&utm_campaign=partner_m_webdealer_ooo_online_services_t2-ru_online",
            final_url="https://spb.t2.ru/promo/partner-all?utm_campaign=partner_m_webdealer_ooo_online_services_t2-ru_online&utm_medium=t2-ru_online&utm_source=webdealer&pageParams=askForRegion%3Dtrue",
        ),
        index,
    )

    assert result.status == VALIDATION_STATUS_OK
    assert result.match_key == "t2-ru.online::\u043f\u0430\u0440\u0442\u043d\u0435\u0440 \u043c"
    assert result.reference_part == "t2.ru/promo/partner-all?utm_campaign=partner_m_webdealer_ooo_online_services_t2-ru_online&utm_medium=t2-ru_online&utm_source=webdealer&pageParams=askForRegion%3Dtrue"


def test_reference_loader_treats_root_page_url_as_fallback(tmp_path: Path) -> None:
    reference_path = tmp_path / "Links_mobile_tarriffs.xlsx"
    _write_xlsx(
        reference_path,
        REFERENCE_HEADERS,
        [
            [
                "mts-home-gpon.ru",
                "https://mts-home-gpon.ru/",
                "МТС Супер",
                "mts.ru/personal/mobilnaya-svyaz/tarifi/vse-tarifi/mts-super/?utm_source=mtspn",
                "contains",
                "",
                True,
            ],
        ],
    )

    references = load_reference_links(reference_path)
    index = ReferenceIndex.build(references)

    result = evaluate_validation(
        _build_input_row(
            domain="mts-home-gpon.ru",
            checked_page_url="https://mts-home-gpon.ru/moskva/mobilnaya-svyaz",
            tariff_name="МТС Супер",
            click_url="https://moskva.mts.ru/personal/mobilnaya-svyaz/tarifi/vse-tarifi/mts-super/?utm_source=mtspn&utm_medium=cpa",
            final_url="https://moskva.mts.ru/personal/mobilnaya-svyaz/tarifi/vse-tarifi/mts-super/?utm_source=mtspn&utm_medium=cpa",
        ),
        index,
    )

    assert result.status == VALIDATION_STATUS_OK
    assert result.reference_part == "mts.ru/personal/mobilnaya-svyaz/tarifi/vse-tarifi/mts-super/?utm_source=mtspn"


def test_validation_handles_regex_and_fallback(tmp_path: Path) -> None:
    reference_path = tmp_path / "Links_mobile_tarriffs.xlsx"
    _write_xlsx(
        reference_path,
        REFERENCE_HEADERS,
        [
            [
                "t2-ru.online",
                "",
                "ПАРТНЕР М",
                r"t2\.ru/promo/partner-all\?utm_source=webdealer",
                "regex",
                "",
                True,
            ],
        ],
    )

    references = load_reference_links(reference_path)
    index = ReferenceIndex.build(references)

    regex_result = evaluate_validation(
        _build_input_row(
            tariff_name="ПАРТНЕР М",
            click_url="https://t2.ru/promo/partner-all?utm_source=webdealer&utm_medium=t2-ru_online",
        ),
        index,
    )
    assert regex_result.status == VALIDATION_STATUS_OK

    fallback_result = evaluate_validation(
        _build_input_row(
            tariff_name="ПАРТНЕР М",
            click_url="",
            final_url="https://t2.ru/promo/partner-all?utm_source=webdealer&utm_medium=t2-ru_online",
        ),
        index,
        use_final_url_as_fallback=True,
    )
    assert fallback_result.status == VALIDATION_STATUS_OK


def test_validation_statuses_for_missing_reference_and_missing_link() -> None:
    index = ReferenceIndex.build([])

    no_fact_result = evaluate_validation(
        _build_input_row(click_url="", final_url=""),
        index,
    )
    assert no_fact_result.status == VALIDATION_STATUS_NO_FACTUAL_LINK

    no_ref_result = evaluate_validation(
        _build_input_row(click_url="https://example.com/landing"),
        index,
    )
    assert no_ref_result.status == VALIDATION_STATUS_NO_REFERENCE


def test_validation_report_writer_and_summary(tmp_path: Path) -> None:
    reference_path = tmp_path / "Links_mobile_tarriffs.xlsx"
    _write_xlsx(
        reference_path,
        REFERENCE_HEADERS,
        [
            [
                "t2-ru.online",
                "https://t2-ru.online/mobilnaya-svyaz",
                "МОЙ ОНЛАЙН",
                "krasnodar.t2.ru/tariffs",
                "contains",
                "",
                True,
            ],
        ],
    )
    references = load_reference_links(reference_path)
    index = ReferenceIndex.build(references)

    rows = [
        evaluate_validation(_build_input_row(tariff_name="МОЙ ОНЛАЙН"), index).row,
        evaluate_validation(_build_input_row(tariff_name="МОЙ ОНЛАЙН", click_url="https://example.com/other"), index).row,
        evaluate_validation(_build_input_row(tariff_name="BLACK", click_url="https://example.com/other"), index).row,
        evaluate_validation(_build_input_row(tariff_name="PREMIUM", click_url=""), index).row,
    ]

    report_path = write_validation_report(rows, tmp_path / "validated.xlsx")
    assert report_path.exists()

    workbook = load_workbook(report_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [cell.value for cell in sheet[1]]
    assert "Эталонная часть ссылки" in headers
    assert "Итог проверки партнерской ссылки" in headers
    assert "Ключ сопоставления" in headers

    summary = summarize_validation_rows(rows, str(report_path), "pilot")
    assert summary.ok_rows == 1
    assert summary.not_ok_rows == 1
    assert summary.no_reference_rows == 1
    assert summary.no_factual_link_rows == 1
    assert summary.product_error_rows == 2


def test_validation_alert_message_includes_artifact_link() -> None:
    summary = summarize_validation_rows(
        [
            replace(_build_input_row(), validation_status=VALIDATION_STATUS_OK),
            replace(_build_input_row(tariff_name="BLACK"), validation_status=VALIDATION_STATUS_NOT_OK),
        ],
        "reports/partner_links_mobile_validated_123.xlsx",
        "pilot",
    )

    message = build_validation_alert_message(
        summary,
        checked_at="2026-06-11 09:55:23",
        build_url="https://jenkins.example/job/1/",
        report_path="reports/partner_links_mobile_validated_123.xlsx",
    )

    assert message == (
        "Партнерские ссылки на лендах\n"
        "\n"
        "Дата проверки: 11.06.2026 09:55\n"
        "\n"
        "Всего проверено тарифов: 2\n"
        "\n"
        "OK: 1\n"
        "НЕ OK: 1\n"
        "Нет эталона: 0\n"
        "Нет фактической ссылки: 0\n"
        "\n"
        "Отчет: https://jenkins.example/job/1/artifact/reports/partner_links_mobile_validated_123.xlsx"
    )


def test_alert_sender_skips_without_proxy_env(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.delenv("TELEGRAM_PROXY_URL", raising=False)
    monkeypatch.delenv("TELEGRAM_PROXY_AUTH_SECRET", raising=False)
    monkeypatch.delenv("TELEGRAM_PROXY_CHAT_CREDENTIAL", raising=False)
    monkeypatch.delenv("TELEGRAM_PROXY_CREDS", raising=False)

    summary = summarize_validation_rows(
        [
            replace(_build_input_row(), validation_status=VALIDATION_STATUS_OK),
            replace(_build_input_row(tariff_name="BLACK"), validation_status=VALIDATION_STATUS_NOT_OK),
            replace(_build_input_row(tariff_name="PREMIUM"), validation_status=VALIDATION_STATUS_NO_REFERENCE),
        ],
        "reports/validated.xlsx",
        "pilot",
    )
    message = build_validation_alert_message(summary, "2026-06-11 09:55:23", "https://jenkins.example/job/1/")
    result = send_validation_alert(message)
    assert result.status == "skipped"


def test_alert_sender_uses_proxy_payload(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("TELEGRAM_PROXY_URL", "https://proxy.example/telegram")
    monkeypatch.setenv("TELEGRAM_PROXY_AUTH_SECRET", "proxy-secret")
    monkeypatch.setenv("TELEGRAM_PROXY_CREDS", "chat-creds")

    captured: dict[str, object] = {}

    class FakeResponse:
        status = 200

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

        def read(self) -> bytes:
            return b"ok"

    def fake_urlopen(request, timeout=0):
        captured["request"] = request
        captured["timeout"] = timeout
        return FakeResponse()

    monkeypatch.setattr("src.alert_sender.urlrequest.urlopen", fake_urlopen)

    result = send_validation_alert("Отчет готов")

    assert result.sent is True
    assert result.status == "sent"
    assert captured["timeout"] == 20

    request = captured["request"]
    assert request.full_url == "https://proxy.example/telegram"
    assert request.get_method() == "POST"

    headers = {key.lower(): value for key, value in request.header_items()}
    assert headers["x-authentication"] == "proxy-secret"
    assert headers["content-type"] == "application/json"

    payload = json.loads(request.data.decode("utf-8"))
    assert payload == {
        "title": "Partner_links_mobile [summary]",
        "text": "Отчет готов",
        "creds": "chat-creds",
        "parse_mode": "HTML",
        "disable_notification": False,
    }


def test_input_loader_uses_latest_report_from_reports_dir(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.chdir(tmp_path)
    reports_dir = tmp_path / "reports"
    reports_dir.mkdir()

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "Домен",
            "URL проверяемой страницы",
            "Название тарифа",
            "Ссылка после клика",
            "Статус загрузки страницы",
            "HTTP-статус",
            "Финальный URL",
            "Ошибка",
            "Дата и время проверки",
            "Оператор",
            "Номер карточки",
            "Тип перехода",
            "Исходный href",
            "Время загрузки, мс",
            "Окружение запуска",
            "Признак продуктовой ошибки",
        ]
    )
    sheet.append(
        [
            "old.example",
            "https://old.example/page",
            "OLD",
            "https://old.example/click",
            "Успешно",
            "200",
            "https://old.example/final",
            "",
            "2026-06-17 11:35:00",
            "OLD",
            "1",
            "new_tab",
            "https://old.example/href",
            "100",
            "run_mode=pilot",
            False,
        ]
    )
    old_report = reports_dir / "partner_links_mobile_2026-06-17_11-35.xlsx"
    workbook.save(old_report)

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "Домен",
            "URL проверяемой страницы",
            "Название тарифа",
            "Ссылка после клика",
            "Статус загрузки страницы",
            "HTTP-статус",
            "Финальный URL",
            "Ошибка",
            "Дата и время проверки",
            "Оператор",
            "Номер карточки",
            "Тип перехода",
            "Исходный href",
            "Время загрузки, мс",
            "Окружение запуска",
            "Признак продуктовой ошибки",
        ]
    )
    sheet.append(
        [
            "new.example",
            "https://new.example/page",
            "NEW",
            "https://new.example/click",
            "Успешно",
            "200",
            "https://new.example/final",
            "",
            "2026-06-17 11:40:00",
            "NEW",
            "2",
            "new_tab",
            "https://new.example/href",
            "100",
            "run_mode=pilot",
            False,
        ]
    )
    latest_report = reports_dir / "partner_links_mobile_2026-06-17_11-40.xlsx"
    workbook.save(latest_report)

    older_mtime = 1_750_000_000
    latest_mtime = 1_750_000_600
    latest_report.touch()
    old_report.touch()
    import os

    os.utime(old_report, (older_mtime, older_mtime))
    os.utime(latest_report, (latest_mtime, latest_mtime))

    rows = load_input_rows("")
    assert len(rows) == 1
    assert rows[0].domain == "new.example"
